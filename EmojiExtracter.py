import emoji
import re
import openpyxl
import pandas as pd


data = []
#Reads in data from excel
fname = 'Spreadsheet.xlsx'
wb = openpyxl.load_workbook(fname)
sheet = wb.get_sheet_by_name('50aTime')

for rowOfCellObjects in sheet['A1':'A18956']:
  for cellObj in rowOfCellObjects:
    #Iterate through each cell in the range
    line = cellObj.value
    #regex to grab only emojis from captions
    regex = re.compile(r'(\u00a9|\u00ae|[\u2050-\u2099]|[\u2601-\u2605]|[\u2610-\u2660]|[\u2662-\u3000]|\ud83c[\ud000-\udfff]|\ud83d[\ud000-\udfff]|\ud83e[\ud000-\udfff])')
    onlyEmojis = regex.findall(line)
    #Convert emojis to text
    textonlyEmojis = emoji.demojize(onlyEmojis)
    #Regex to get rid of colons and seperate out emoji names into seperate cells
    final = re.findall(":+([^@][^#][^:]+):", textonlyEmojis)
    data.append(final)

# raw_data = emoji.demojize("🤔 🙈 me así, se 😌 ds 💕👭👙 hello 👩🏾‍🎓 emoji hello 👨‍👩‍👦‍👦 how are 😊 you today🙅🏽🙅🏽")
# onlyEmojis = re.findall(":+([^:]+):",raw_data)
#REGEX to extract emojis
#:+([^:]+):
#:+([^@][^#][^:]+):

# for i in data:
#   print(i)

# wb = openpyxl.Workbook()
# sheet = wb.active

#
# for i in range(1,20):
#   if (data[(i-1)] != []):
#     sheet.cell(row = i, column = 8).value = data[(i-1)][0]
#exports to excel
df = pd.DataFrame(data)
with pd.ExcelWriter('Spreadsheet(COPY).xlsx', mode='a') as writer:
  df.to_excel(writer)
